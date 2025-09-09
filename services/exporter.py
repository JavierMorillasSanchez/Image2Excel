# services/exporter.py
"""
Exportador a Excel sencillo con openpyxl.

API esperada por image2excel.adapters.OpenpyxlExporterAdapter:
    class ExcelExporter:
        def export_table(self, rows: list[list[str]], output_dir: str, filename: str) -> str
"""

from __future__ import annotations
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self, sheet_name: str = "Texto Extraído") -> None:
        self.sheet_name = sheet_name

    def export_table(self, rows: List[List[str]], output_dir: str, filename: str) -> str:
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / filename

        # Normalizar filas a listas de strings
        normalized_rows: List[List[str]] = []
        for r in rows or [[]]:
            if not isinstance(r, (list, tuple)):
                normalized_rows.append([str(r)])
            else:
                normalized_rows.append(["" if v is None else str(v) for v in r])

        # Crear rejilla rectangular: padding con "" hasta el número máximo de columnas
        max_cols = max((len(r) for r in normalized_rows), default=0)
        if max_cols == 0:
            normalized_rows = [[""]]
            max_cols = 1
        for r in normalized_rows:
            if len(r) < max_cols:
                r.extend([""] * (max_cols - len(r)))

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        for r in normalized_rows:
            ws.append(r)

        # Bordes finos en todas las celdas
        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=len(normalized_rows), min_col=1, max_col=max_cols):
            for cell in row:
                cell.border = thin_border

        # Auto-ajuste de ancho de columnas según texto más largo
        for col_idx in range(1, max_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, len(normalized_rows) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value)
                if len(text) > max_len:
                    max_len = len(text)
            # Pequeño padding para mejor visualización
            ws.column_dimensions[col_letter].width = max(8, min(60, max_len + 2))

        wb.save(out_path)
        return str(out_path)
