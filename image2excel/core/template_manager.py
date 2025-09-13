from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import json
from typing import List
import cv2
import numpy as np

@dataclass(frozen=True)
class Cell:
    row: int
    col: int
    # Si normalized=True => x,y,w,h ∈ [0..1]; si False => píxeles absolutos
    x: float
    y: float
    w: float
    h: float
    name: str | None = None

@dataclass(frozen=True)
class GridSpec:
    name: str
    width: int         # tamaño CANÓNICO al que alineamos (p.ej. 1280x720)
    height: int
    cells: List[Cell]
    normalized: bool = True  # <- por defecto trabajamos en relativo

class TemplateManager:
    def __init__(self, templates_dir: str = "templates"):
        self._dir = Path(templates_dir)
        self._dir.mkdir(parents=True, exist_ok=True)

    def save(self, spec: GridSpec) -> Path:
        path = self._dir / f"{spec.name}.json"
        payload = {
            "name": spec.name,
            "width": spec.width,
            "height": spec.height,
            "normalized": spec.normalized,
            "cells": [cell.__dict__ for cell in spec.cells],
        }
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return path

    def load(self, name: str) -> GridSpec:
        path = self._dir / f"{name}.json"

        # Si no existe la plantilla, crear una por defecto
        if not path.exists():
            if name == "default_template":
                return self._create_default_template()
            else:
                raise FileNotFoundError(f"Plantilla no encontrada: {path}")

        data = json.loads(path.read_text(encoding="utf-8"))
        normalized = bool(data.get("normalized", False))  # retro-compat
        cells = [Cell(**c) for c in data["cells"]]
        return GridSpec(
            name=data["name"],
            width=int(data["width"]),
            height=int(data["height"]),
            cells=cells,
            normalized=normalized,
        )

    def _create_default_template(self) -> GridSpec:
        """Crea una plantilla por defecto para pruebas."""
        # Plantilla básica de ejemplo con coordenadas normalizadas
        cells = [
            Cell(row=1, col=1, x=0.1, y=0.1, w=0.2, h=0.15, name="Etiqueta"),
            Cell(row=1, col=2, x=0.35, y=0.1, w=0.2, h=0.15, name="DXO"),
            Cell(row=1, col=3, x=0.6, y=0.1, w=0.2, h=0.15, name="Marca"),
            Cell(row=1, col=4, x=0.85, y=0.1, w=0.1, h=0.15, name="Peso"),
            # Fila 2
            Cell(row=2, col=1, x=0.1, y=0.3, w=0.2, h=0.15, name="Etiqueta"),
            Cell(row=2, col=2, x=0.35, y=0.3, w=0.2, h=0.15, name="DXO"),
            Cell(row=2, col=3, x=0.6, y=0.3, w=0.2, h=0.15, name="Marca"),
            Cell(row=2, col=4, x=0.85, y=0.3, w=0.1, h=0.15, name="Peso"),
            # Fila 3
            Cell(row=3, col=1, x=0.1, y=0.5, w=0.2, h=0.15, name="Etiqueta"),
            Cell(row=3, col=2, x=0.35, y=0.5, w=0.2, h=0.15, name="DXO"),
            Cell(row=3, col=3, x=0.6, y=0.5, w=0.2, h=0.15, name="Marca"),
            Cell(row=3, col=4, x=0.85, y=0.5, w=0.1, h=0.15, name="Peso"),
        ]

        spec = GridSpec(
            name="default_template",
            width=1280,
            height=720,
            cells=cells,
            normalized=True
        )

        # Guardar la plantilla por defecto para futuros usos
        self.save(spec)

        return spec

    @staticmethod
    def image_from_any(path: str) -> np.ndarray:
        """Carga una imagen desde cualquier formato soportado."""
        path_obj = Path(path)

        if path_obj.suffix.lower() in ['.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff']:
            # Imagen directa
            img = cv2.imread(str(path_obj))
            if img is None:
                raise ValueError(f"No se pudo cargar la imagen: {path_obj}")
            return img
        elif path_obj.suffix.lower() == '.pdf':
            # PDF - convertir primera página a imagen
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(str(path_obj))
                page = doc[0]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom
                img_data = pix.tobytes("bmp")
                img = cv2.imdecode(np.frombuffer(img_data, np.uint8), cv2.IMREAD_COLOR)
                doc.close()
                return img
            except ImportError:
                raise ImportError("PyMuPDF no está instalado. Instale con: pip install PyMuPDF")
        elif path_obj.suffix.lower() in ['.xlsx', '.xls']:
            # Excel - convertir a imagen
            try:
                return TemplateManager._excel_to_image(str(path_obj))
            except ImportError:
                raise ImportError("openpyxl no está instalado. Instale con: pip install openpyxl")
        elif path_obj.suffix.lower() == '.docx':
            # Word - convertir a imagen
            try:
                return TemplateManager._docx_to_image(str(path_obj))
            except ImportError:
                raise ImportError("python-docx no está instalado. Instale con: pip install python-docx")
        else:
            raise ValueError(f"Formato no soportado: {path_obj.suffix}")

    @staticmethod
    def _excel_to_image(excel_path: str) -> np.ndarray:
        """Convierte un archivo Excel a imagen."""
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            import io

            # Cargar el archivo Excel
            wb = load_workbook(excel_path)
            ws = wb.active

            # Crear una imagen simple basada en el contenido
            # Para simplificar, creamos una imagen con el texto de las celdas
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError("El archivo Excel está vacío")

            # Calcular dimensiones aproximadas
            max_cols = max(len(row) for row in rows if row)
            num_rows = len([row for row in rows if any(cell for cell in row)])

            # Crear imagen base
            cell_width = 100
            cell_height = 30
            img_width = max_cols * cell_width
            img_height = num_rows * cell_height

            # Crear imagen blanca
            img = np.ones((img_height, img_width, 3), dtype=np.uint8) * 255

            # Dibujar líneas de cuadrícula
            for i in range(num_rows + 1):
                y = i * cell_height
                cv2.line(img, (0, y), (img_width, y), (0, 0, 0), 1)

            for i in range(max_cols + 1):
                x = i * cell_width
                cv2.line(img, (x, 0), (x, img_height), (0, 0, 0), 1)

            return img

        except Exception as e:
            raise ValueError(f"Error al convertir Excel a imagen: {e}")

    @staticmethod
    def _docx_to_image(docx_path: str) -> np.ndarray:
        """Convierte un archivo Word a imagen."""
        try:
            from docx import Document

            # Cargar el documento
            doc = Document(docx_path)

            # Crear imagen simple basada en el contenido
            # Para simplificar, creamos una imagen con texto
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

            if not paragraphs:
                raise ValueError("El documento Word está vacío")

            # Crear imagen base
            img_width = 800
            img_height = len(paragraphs) * 40 + 100

            img = np.ones((img_height, img_width, 3), dtype=np.uint8) * 255

            # Dibujar líneas horizontales para simular texto
            for i, _ in enumerate(paragraphs):
                y = 50 + i * 40
                cv2.line(img, (50, y), (img_width - 50, y), (0, 0, 0), 1)

            return img

        except Exception as e:
            raise ValueError(f"Error al convertir Word a imagen: {e}")
