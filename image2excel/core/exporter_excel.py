"""
Exportador de Excel para Image2Excel.

Exporta datos estructurados a archivos Excel con formato.
"""

from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_rows_xlsx(headers: List[str], rows: List[Dict[str, Any]], output_path: str) -> str:
    """
    Exporta filas de datos a un archivo Excel con formato.

    Args:
        headers: Lista de nombres de columnas
        rows: Lista de diccionarios con los datos de cada fila
        output_path: Ruta del archivo Excel de salida

    Returns:
        Ruta del archivo generado
    """
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Extraídos"

    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    cell_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir datos
    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Ajustar ancho de columnas
    for col, header in enumerate(headers, 1):
        column_letter = get_column_letter(col)

        # Calcular ancho basado en el contenido
        max_length = len(header)
        for row in rows:
            value = str(row.get(header, ""))
            max_length = max(max_length, len(value))

        # Establecer ancho con un mínimo y máximo
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Aplicar filtros a los encabezados
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Congelar la primera fila
    ws.freeze_panes = "A2"

    # Guardar archivo
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)

    return str(output_path_obj.resolve())
